from fastapi import FastAPI, APIRouter, HTTPException, File, UploadFile, Request, Response, Cookie
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import PyPDF2
import openpyxl
import io
import httpx

from conference_service import ConferenceRoomService
from research_service import ResearchService
from deliverable_service import DeliverableService
from fast_ai_research import FastAIResearch

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Initialize services
conference_service = ConferenceRoomService()
research_service = ResearchService()
fast_ai_research = FastAIResearch()
deliverable_service = DeliverableService()

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# ==================== AUTH MODELS ====================
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SessionRequest(BaseModel):
    session_id: str


# ==================== PROJECT MODELS ====================
class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    problem_statement: str
    project_type: str
    user_id: Optional[str] = None  # Link to user
    status: str = "in_progress"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    research_data: Optional[Dict[str, Any]] = None
    debate_data: Optional[Dict[str, Any]] = None
    deliverables: List[str] = []

class ProjectCreate(BaseModel):
    title: str
    problem_statement: str
    project_type: str

class ResearchRequest(BaseModel):
    project_id: str
    vendor_name: Optional[str] = None
    industry: Optional[str] = None
    query: Optional[str] = None
    additional_context: Optional[str] = None

class DebateRequest(BaseModel):
    project_id: str
    problem: str

class DeliverableRequest(BaseModel):
    project_id: str
    deliverable_type: str
    content: Dict[str, Any]

class APIKeyUpdate(BaseModel):
    key_name: str
    key_value: str


# ==================== AUTH HELPER ====================
async def get_current_user(request: Request) -> Optional[User]:
    """Get current user from session token cookie or Authorization header"""
    session_token = request.cookies.get("session_token")
    
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header[7:]
    
    if not session_token:
        return None
    
    session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session:
        return None
    
    # Check expiry with timezone awareness
    expires_at = session.get("expires_at")
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        return None
    
    user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user:
        return None
    
    return User(**user)


# ==================== AUTH ENDPOINTS ====================
@api_router.post("/auth/session")
async def create_session(request: SessionRequest, response: Response):
    """Exchange session_id for session_token after Google OAuth"""
    try:
        # Call Emergent auth service to get user data
        async with httpx.AsyncClient() as client:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": request.session_id},
                timeout=10.0
            )
        
        if auth_response.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid session ID")
        
        user_data = auth_response.json()
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        
        # Check if user exists
        existing_user = await db.users.find_one({"email": user_data["email"]}, {"_id": 0})
        if existing_user:
            user_id = existing_user["user_id"]
        else:
            # Create new user
            new_user = {
                "user_id": user_id,
                "email": user_data["email"],
                "name": user_data["name"],
                "picture": user_data.get("picture", ""),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(new_user)
        
        # Create session
        session_token = user_data.get("session_token", f"session_{uuid.uuid4().hex}")
        expires_at = datetime.now(timezone.utc) + timedelta(days=7)
        
        await db.user_sessions.insert_one({
            "user_id": user_id,
            "session_token": session_token,
            "expires_at": expires_at.isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        # Set cookie
        response.set_cookie(
            key="session_token",
            value=session_token,
            httponly=True,
            secure=True,
            samesite="none",
            path="/",
            max_age=7 * 24 * 60 * 60  # 7 days
        )
        
        user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
        return user
        
    except httpx.RequestError as e:
        raise HTTPException(status_code=500, detail=f"Auth service error: {str(e)}")

@api_router.get("/auth/me")
async def get_me(request: Request):
    """Get current authenticated user"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user.model_dump()

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    """Logout and clear session"""
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    
    response.delete_cookie(key="session_token", path="/")
    return {"success": True}


# ==================== FILE HELPERS ====================
def extract_text_from_pdf(file_content: bytes) -> str:
    """Extract text from PDF"""
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"Error reading PDF: {str(e)}"

def extract_text_from_excel(file_content: bytes) -> str:
    """Extract text from Excel"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        text = ""
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text += f"\n--- Sheet: {sheet_name} ---\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        return text
    except Exception as e:
        return f"Error reading Excel: {str(e)}"


# ==================== API ROUTES ====================
@api_router.get("/")
async def root():
    return {"message": "Consultant AI Backend API"}

@api_router.get("/agents")
async def get_agents():
    """Get all consultant agents"""
    return conference_service.get_agent_info()

# Project endpoints
@api_router.post("/projects", response_model=Project)
async def create_project(project: ProjectCreate, request: Request):
    """Create a new consulting project"""
    user = await get_current_user(request)
    
    project_dict = project.model_dump()
    project_obj = Project(**project_dict)
    
    # Link to user if authenticated
    if user:
        project_obj.user_id = user.user_id
    
    doc = project_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.projects.insert_one(doc)
    return project_obj

@api_router.get("/projects", response_model=List[Project])
async def get_projects(request: Request):
    """Get all projects (filtered by user if authenticated)"""
    user = await get_current_user(request)
    
    query = {}
    if user:
        query["user_id"] = user.user_id
    
    projects = await db.projects.find(query, {"_id": 0}).to_list(1000)
    
    for project in projects:
        if isinstance(project.get('created_at'), str):
            project['created_at'] = datetime.fromisoformat(project['created_at'])
        if isinstance(project.get('updated_at'), str):
            project['updated_at'] = datetime.fromisoformat(project['updated_at'])
    
    return projects

@api_router.get("/projects/{project_id}", response_model=Project)
async def get_project(project_id: str):
    """Get a specific project"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    if isinstance(project.get('created_at'), str):
        project['created_at'] = datetime.fromisoformat(project['created_at'])
    if isinstance(project.get('updated_at'), str):
        project['updated_at'] = datetime.fromisoformat(project['updated_at'])
    
    return project

# File upload endpoint
@api_router.post("/research/upload")
async def upload_research_files(files: List[UploadFile] = File(...)):
    """Upload PDF/Excel files for additional context"""
    extracted_content = []
    
    for file in files:
        file_content = await file.read()
        
        if file.filename.endswith('.pdf'):
            text = extract_text_from_pdf(file_content)
            extracted_content.append({
                "filename": file.filename,
                "type": "pdf",
                "content": text[:5000]
            })
        elif file.filename.endswith(('.xlsx', '.xls')):
            text = extract_text_from_excel(file_content)
            extracted_content.append({
                "filename": file.filename,
                "type": "excel",
                "content": text[:5000]
            })
        else:
            extracted_content.append({
                "filename": file.filename,
                "type": "unsupported",
                "content": "File type not supported. Please upload PDF or Excel files."
            })
    
    return {
        "success": True,
        "files_processed": len(extracted_content),
        "extracted_content": extracted_content
    }

# Research endpoints
@api_router.post("/research/vendor-analysis")
async def conduct_vendor_analysis(request_data: ResearchRequest, request: Request):
    """Conduct fast AI-powered vendor analysis"""
    user = await get_current_user(request)
    
    problem = request_data.query or "General business analysis"
    
    results = await fast_ai_research.research(
        problem=problem,
        vendor_name=request_data.vendor_name,
        industry=request_data.industry,
        additional_context=request_data.additional_context
    )
    
    # Update project with research data
    update_data = {
        "research_data": results, 
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if user:
        update_data["user_id"] = user.user_id
        
    await db.projects.update_one(
        {"id": request_data.project_id},
        {"$set": update_data}
    )
    
    return results

@api_router.post("/research/search")
async def search_information(request_data: ResearchRequest):
    """Search for public information"""
    if not request_data.query:
        raise HTTPException(status_code=400, detail="query is required")
    
    results = await research_service.search_public_information(request_data.query)
    return results

# Conference room endpoints
@api_router.post("/conference/debate")
async def conduct_debate(request_data: DebateRequest, request: Request):
    """Conduct a multi-agent debate"""
    user = await get_current_user(request)
    
    results = await conference_service.conduct_debate(request_data.problem)
    
    update_data = {
        "debate_data": results, 
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if user:
        update_data["user_id"] = user.user_id
    
    await db.projects.update_one(
        {"id": request_data.project_id},
        {"$set": update_data}
    )
    
    return results

# Deliverable endpoints
@api_router.post("/deliverables/excel")
async def generate_excel(request_data: DeliverableRequest):
    """Generate Excel deliverable"""
    filename = deliverable_service.generate_excel_report(
        request_data.content,
        request_data.deliverable_type
    )
    
    await db.projects.update_one(
        {"id": request_data.project_id},
        {
            "$push": {"deliverables": filename},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        }
    )
    
    return {"filename": filename, "type": "excel"}

@api_router.post("/deliverables/ppt")
async def generate_ppt(request_data: DeliverableRequest):
    """Generate professional PowerPoint presentation using Gamma AI"""
    # Fetch project data to include debate and research data
    project = await db.projects.find_one({"id": request_data.project_id})

    content = request_data.content.copy()
    if project:
        content['debate_data'] = project.get('debate_data', {})
        content['research_data'] = project.get('research_data', {})

    # Use Gamma API for professional presentations (falls back to python-pptx)
    result = await deliverable_service.generate_ppt_via_gamma(content)

    if result.get('success'):
        # For Gamma presentations, store the URL; for local, store filename
        deliverable_info = result.get('presentation_url') or result.get('filename', 'presentation.pptx')

        await db.projects.update_one(
            {"id": request_data.project_id},
            {
                "$push": {"deliverables": deliverable_info},
                "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
            }
        )

    return result

@api_router.post("/deliverables/pdf")
async def generate_pdf(request_data: DeliverableRequest):
    """Generate professional PDF report"""
    # Fetch project data to include debate and research data
    project = await db.projects.find_one({"id": request_data.project_id})

    content = request_data.content.copy()
    if project:
        content['debate_data'] = project.get('debate_data', {})
        content['research_data'] = project.get('research_data', {})

    result = await deliverable_service.generate_pdf(content)

    if result.get('success'):
        filename = result.get('filename', 'report.pdf')

        await db.projects.update_one(
            {"id": request_data.project_id},
            {
                "$push": {"deliverables": filename},
                "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
            }
        )

    return result

@api_router.post("/deliverables/all")
async def generate_all_deliverables(request_data: DeliverableRequest):
    """Generate all deliverables (Excel, PPT, PDF) for a project"""
    # Fetch project data
    project = await db.projects.find_one({"id": request_data.project_id})

    content = request_data.content.copy()
    if project:
        content['debate_data'] = project.get('debate_data', {})
        content['research_data'] = project.get('research_data', {})

    results = {
        "excel": None,
        "ppt": None,
        "pdf": None,
        "success": True
    }

    # Generate Excel
    try:
        excel_filename = deliverable_service.generate_excel_report(content, request_data.deliverable_type)
        results["excel"] = {"filename": excel_filename, "success": True}
    except Exception as e:
        results["excel"] = {"success": False, "error": str(e)}

    # Generate PPT using Gamma AI
    try:
        ppt_result = await deliverable_service.generate_ppt_via_gamma(content)
        results["ppt"] = ppt_result
    except Exception as e:
        results["ppt"] = {"success": False, "error": str(e)}

    # Generate PDF
    try:
        pdf_result = await deliverable_service.generate_pdf(content)
        results["pdf"] = pdf_result
    except Exception as e:
        results["pdf"] = {"success": False, "error": str(e)}

    # Update project with all deliverables
    deliverable_filenames = []
    if results["excel"] and results["excel"].get("success"):
        deliverable_filenames.append(results["excel"]["filename"])
    if results["ppt"] and results["ppt"].get("success"):
        # Handle both Gamma URLs and local filenames
        ppt_deliverable = results["ppt"].get("presentation_url") or results["ppt"].get("filename")
        if ppt_deliverable:
            deliverable_filenames.append(ppt_deliverable)
    if results["pdf"] and results["pdf"].get("success"):
        deliverable_filenames.append(results["pdf"].get("filename"))

    if deliverable_filenames:
        await db.projects.update_one(
            {"id": request_data.project_id},
            {
                "$push": {"deliverables": {"$each": deliverable_filenames}},
                "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
            }
        )

    return results

@api_router.get("/deliverables/{filename}")
async def download_deliverable(filename: str):
    """Download a deliverable file"""
    # Try multiple possible paths
    possible_paths = [
        f"/home/user/aiconsultant/deliverables/{filename}",
        f"/app/deliverables/{filename}"
    ]

    filepath = None
    for path in possible_paths:
        if os.path.exists(path):
            filepath = path
            break

    if not filepath:
        raise HTTPException(status_code=404, detail="File not found")

    # Determine media type based on extension
    media_types = {
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
        '.pdf': 'application/pdf',
        '.html': 'text/html'
    }

    ext = os.path.splitext(filename)[1].lower()
    media_type = media_types.get(ext, 'application/octet-stream')

    return FileResponse(
        filepath,
        filename=filename,
        media_type=media_type
    )

# Get all deliverables for listing
@api_router.get("/deliverables")
async def list_deliverables():
    """List all available deliverable files"""
    possible_dirs = [
        "/home/user/aiconsultant/deliverables",
        "/app/deliverables"
    ]

    deliverables_dir = None
    for dir_path in possible_dirs:
        if os.path.exists(dir_path):
            deliverables_dir = dir_path
            break

    if not deliverables_dir:
        return {"files": []}

    files = []
    for f in os.listdir(deliverables_dir):
        filepath = os.path.join(deliverables_dir, f)
        if os.path.isfile(filepath):
            ext = os.path.splitext(f)[1].lower()
            file_type = {
                '.xlsx': 'Excel',
                '.pptx': 'PowerPoint',
                '.pdf': 'PDF',
                '.html': 'HTML',
                '.txt': 'Text'
            }.get(ext, 'Unknown')

            files.append({
                "filename": f,
                "size": os.path.getsize(filepath),
                "type": file_type,
                "created": datetime.fromtimestamp(os.path.getctime(filepath)).isoformat()
            })

    return {"files": sorted(files, key=lambda x: x['created'], reverse=True)}

# Settings endpoints
@api_router.post("/settings/api-keys")
async def update_api_key(key_update: APIKeyUpdate):
    """Update API key (stored in environment for session)"""
    os.environ[key_update.key_name] = key_update.key_value
    return {"success": True, "message": f"API key {key_update.key_name} updated"}

# Mount the router
app.include_router(api_router)

# CORS middleware - handle credentials properly
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "https://strategyai-10.preview.emergentagent.com",
        "https://auth.emergentagent.com"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
